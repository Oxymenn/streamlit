import streamlit as st
import pandas as pd
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import ast
import openpyxl
from openpyxl.styles import PatternFill

def load_file(uploaded_file):
    file_type = uploaded_file.name.split('.')[-1]
    if file_type == 'xlsx':
        df = pd.read_excel(uploaded_file, engine='openpyxl')
    elif file_type == 'csv':
        df = pd.read_csv(uploaded_file)
    return df

def preprocess_embeddings(df, embedding_col):
    # Convertir les chaînes de caractères en listes de nombres si nécessaire
    if isinstance(df[embedding_col].iloc[0], str):
        df[embedding_col] = df[embedding_col].apply(ast.literal_eval)
    return np.array(df[embedding_col].tolist())

def calculate_cosine_similarity(embeddings):
    similarities = cosine_similarity(embeddings)
    return similarities

def generate_similarity_table(df, url_column, similarities, num_links):
    similarity_data = []
    url_to_dest = {url: [] for url in df[url_column]}
    for index, row in df.iterrows():
        similarity_scores = similarities[index]
        similar_indices = np.argsort(similarity_scores)[::-1]
        similar_urls = []
        similar_scores = []
        count = 0
        for idx in similar_indices:
            if df[url_column].iloc[idx] != row[url_column] and count < num_links:
                similar_urls.append(df[url_column].iloc[idx])
                similar_scores.append(similarity_scores[idx])
                count += 1
        url_to_dest[row[url_column]] = similar_urls
        similarity_data.append({
            "URL de départ": row[url_column],
            "URLs de destination": ", ".join(similar_urls),
            "Scores de similarité": ", ".join(map(str, similar_scores))
        })
    similarity_df = pd.DataFrame(similarity_data)
    return similarity_df, url_to_dest

def close_loop(url_to_dest):
    # Create a dictionary to count incoming links
    url_incoming_links = {url: 0 for url in url_to_dest.keys()}

    # Increment the incoming links count
    for dest_urls in url_to_dest.values():
        for dest in dest_urls:
            if dest in url_incoming_links:
                url_incoming_links[dest] += 1

    # Balance the links to ensure each URL gets the same number of links
    balanced_dest = {url: [] for url in url_to_dest.keys()}
    for url, dest_urls in url_to_dest.items():
        for dest in dest_urls:
            if len(balanced_dest[dest]) < 4:  # Ensure each URL gets max 4 incoming links
                balanced_dest[url].append(dest)

    return balanced_dest

def apply_color(sheet, cell, score):
    if 0.9 <= score <= 1:
        fill = PatternFill(start_color='149414', end_color='149414', fill_type='solid')
    elif 0.8 <= score < 0.9:
        fill = PatternFill(start_color='16B84E', end_color='16B84E', fill_type='solid')
    elif 0.75 <= score < 0.8:
        fill = PatternFill(start_color='B0F2B6', end_color='B0F2B6', fill_type='solid')
    elif 0.6 <= score < 0.75:
        fill = PatternFill(start_color='ff4c4c', end_color='ff4c4c', fill_type='solid')
    else:
        fill = PatternFill(start_color='e50000', end_color='e50000', fill_type='solid')
    sheet[cell].fill = fill

def save_to_excel(similarity_table):
    file_name = 'similarity_table.xlsx'
    similarity_table.to_excel(file_name, index=False, engine='openpyxl')
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=3).value
        if isinstance(cell_value, str):
            scores = cell_value.split(", ")
            for score in scores:
                score = float(score)
                cell = f'C{row}'
                apply_color(sheet, cell, score)

    wb.save(file_name)
    return file_name

def app():
    st.title("Analyse de Similarité Cosinus des URL")
    uploaded_file = st.file_uploader("Choisissez un fichier Excel ou CSV", type=["xlsx", "csv"])
    
    if uploaded_file:
        df = load_file(uploaded_file)
        st.write("Aperçu des données :")
        st.write(df.head())
        
        url_column = st.selectbox("Sélectionnez la colonne des URL", df.columns)
        embedding_column = st.selectbox("Sélectionnez la colonne des Embeddings", df.columns)
        
        if st.button("Calculer la similarité cosinus"):
            with st.spinner("Calcul de la similarité en cours..."):
                embeddings = preprocess_embeddings(df, embedding_column)
                similarities = calculate_cosine_similarity(embeddings)
                
                st.session_state.similarities = similarities
                st.session_state.df = df
                st.session_state.url_column = url_column
                st.session_state.embedding_column = embedding_column
                st.session_state.num_links = min(5, len(df))  # 5 liens au lieu de 4
                
                st.write("Calcul de la similarité terminé avec succès !")
                
                similarity_table, url_to_dest = generate_similarity_table(df, url_column, similarities, st.session_state.num_links)
                balanced_dest = close_loop(url_to_dest)

                # Mettre à jour le tableau des similarités avec les liens équilibrés
                similarity_data = []
                for url, dest_urls in balanced_dest.items():
                    similarity_data.append({
                        "URL de départ": url,
                        "URLs de destination": ", ".join(dest_urls)
                    })
                similarity_table = pd.DataFrame(similarity_data)

                st.session_state.similarity_table = similarity_table
                
                st.write("Tableau des similarités :")
                st.write(similarity_table)
                
                excel_file = save_to_excel(similarity_table)
                st.download_button(label="Télécharger le tableau en Excel", data=open(excel_file, 'rb').read(), file_name='similarity_table.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    if 'similarities' in st.session_state:
        df = st.session_state.df
        url_column = st.session_state.url_column
        similarities = st.session_state.similarities
        
        # Curseur pour le nombre de liens à analyser
        num_links = st.slider("Nombre de liens à analyser", min_value=1, max_value=len(df), value=st.session_state.get('num_links', 5))  # 5 liens au lieu de 4
        st.session_state.num_links = num_links
        
        # Sélecteur pour l'URL
        selected_url = st.selectbox("Sélectionnez l'URL pour voir les liens similaires", df[url_column])
        
        if selected_url:
            selected_index = df[df[url_column] == selected_url].index[0]
            similarity_scores = similarities[selected_index]
            similar_indices = np.argsort(similarity_scores)[::-1]
            similar_urls = []
            count = 0
            for idx in similar_indices:
                if df[url_column].iloc[idx] != selected_url and count < num_links:
                    similar_urls.append(df[url_column].iloc[idx])
                    count += 1
            similar_scores = similarity_scores[[df[url_column].iloc[idx] != selected_url for idx in similar_indices]][:num_links]
            
            st.write(f"Top {num_links} URLs les plus similaires à {selected_url} :")
            for url, score in zip(similar_urls, similar_scores):
                st.write(f"{url} (Score: {score})")
            
            # Télécharger les résultats du deuxième rapport en CSV
            report_data = {
                "URL de référence": selected_url,
                "URLs similaires": similar_urls,
                "Scores de similarité": similar_scores
            }
            report_df = pd.DataFrame(report_data)
            report_csv = report_df.to_csv(index=False).encode('utf-8')
            st.download_button(label="Télécharger le rapport en CSV", data=report_csv, file_name=f'similarity_report_{selected_url}.csv', mime='text/csv')

if __name__ == "__main__":
    app()
