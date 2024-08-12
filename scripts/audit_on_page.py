import streamlit as st
import pandas as pd
import cloudscraper
from bs4 import BeautifulSoup
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def analyze_url(url, keyword):
    scraper = cloudscraper.create_scraper()
    try:
        html = scraper.get(url, headers={"User-agent": "Mozilla/5.0"}, timeout=10)
        soup = BeautifulSoup(html.text, 'html.parser')

        metatitle = soup.find('title').get_text() if soup.find('title') else ""
        metadescription = soup.find('meta', attrs={'name':'description'})
        metadescription = metadescription["content"] if metadescription else ""
        h1 = [a.get_text() for a in soup.find_all('h1')]
        h2 = [a.get_text() for a in soup.find_all('h2')]
        paragraph = [a.get_text() for a in soup.find_all('p')]

        metatitle_occurrence = all(term.lower() in metatitle.lower() for term in keyword.split())
        metadescription_occurrence = all(term.lower() in metadescription.lower() for term in keyword.split())
        h1_occurrence = any(all(term.lower() in h.lower() for term in keyword.split()) for h in h1)
        h2_occurrence = any(all(term.lower() in h.lower() for term in keyword.split()) for h in h2)
        paragraph_occurrence = any(all(term.lower() in p.lower() for term in keyword.split()) for p in paragraph)

        return [str(metatitle_occurrence), str(metadescription_occurrence), str(h1_occurrence), str(h2_occurrence), str(paragraph_occurrence)]
    except Exception as e:
        st.error(f"Erreur lors de l'analyse de {url}: {str(e)}")
        return ["Error", "Error", "Error", "Error", "Error"]

def create_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "SEO Analysis"

    for idx, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=idx, value=col)

    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    red_text = Font(color="9C0006")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_text = Font(color="006100")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    dxf_red = DifferentialStyle(font=red_text, fill=red_fill)
    dxf_green = DifferentialStyle(font=green_text, fill=green_fill)

    rule_false = Rule(type="containsText", operator="containsText", text="False", dxf=dxf_red)
    rule_true = Rule(type="containsText", operator="containsText", text="True", dxf=dxf_green)

    for col in ['E', 'F', 'G', 'H', 'I']:  # Colonnes pour les occurrences
        ws.conditional_formatting.add(f'{col}2:{col}{ws.max_row}', rule_false)
        ws.conditional_formatting.add(f'{col}2:{col}{ws.max_row}', rule_true)

    return wb

def app():
    st.title("Analyse SEO On-Site")

    uploaded_file = st.file_uploader("Importer votre fichier Excel", type=['xlsx'])

    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        st.write("Aperçu du fichier importé:")
        st.dataframe(df.head())

        url_column = st.selectbox("Sélectionnez la colonne des URLs", df.columns)
        keyword_column = st.selectbox("Sélectionnez la colonne des mots-clés", df.columns)

        if 'results' not in st.session_state:
            st.session_state.results = None

        if st.button("Exécuter l'analyse") or st.session_state.results is not None:
            if st.session_state.results is None:
                with st.spinner("Analyse en cours..."):
                    results = []
                    progress_bar = st.progress(0)
                    total_rows = len(df)

                    for index, row in df.iterrows():
                        url = row[url_column]
                        keyword = row[keyword_column]
                        result = analyze_url(url, keyword)
                        results.append([url, keyword] + result)
                        progress_bar.progress((index + 1) / total_rows)

                    st.session_state.results = pd.DataFrame(results, columns=["URL", "Keyword", "Metatitle Occurrence", "Metadescription Occurrence", "H1 Occurrence", "H2 Occurrence", "Paragraph Occurrence"])

            st.write("Résultats de l'analyse:")
            st.dataframe(st.session_state.results)

            wb = create_excel(st.session_state.results)
            excel_data = io.BytesIO()
            wb.save(excel_data)
            excel_data.seek(0)

            st.download_button(
                label="Télécharger l'analyse Excel",
                data=excel_data,
                file_name="seo_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    else:
        st.info("Veuillez importer un fichier Excel pour commencer l'analyse.")

if __name__ == "__main__":
    app()
